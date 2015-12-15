# coding: utf-8

from StringIO import StringIO

from django.http.response import HttpResponse
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from rest_framework.exceptions import ParseError
from rest_framework.parsers import FileUploadParser

from pinf.utils import build_exception_response


#################################################
def cellPositions(row,names,dontwant=[]):
    cells = {cell.value: index for index, cell in enumerate(row)}
    for name in dontwant:
        if name:
            if name in cells:
                return None    
    res = []
    for name in names:
        if name:
            if not name in cells:
                return None
            res.append(cells.get(name))
        else:
            res.append(-1)
    return res
         
#################################################
class ExcelParser(FileUploadParser):     
    style1 = []
    style2 = []
    dontwant = []
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    def parse(self, stream, media_type=None, parser_context=None):
        try:
            res = super(ExcelParser, self).parse(stream, media_type, parser_context)
#             backesercizio__pk = parser_context['kwargs']['backesercizio__pk']                        
            file_obj = res.files['file']          
            wb = load_workbook(file_obj, use_iterators = True)
            ws = wb.worksheets[0]
            row = ws.get_squared_range(1, 1, ws.max_column, 1).next()       #get first line with column names
            rows = ws.get_squared_range(1, 2, ws.max_column, ws.max_row)    #get generator for all other rows            
            positions = cellPositions(row,self.style1,self.dontwant)           #try with style1
            if not positions :
                positions = cellPositions(row,self.style2,self.dontwant)            #try with style2
                if not positions :
                    raise ParseError('File non compatibile')                                  
            res.data = []
            for row in rows:
                data = {}
                for index, position in enumerate(positions):                    
                    data[self.style1[index]] = row[position].value if position >= 0 else None
                res.data.append(data)
            return res          
        except Exception, ex:
            raise ParseError(str(ex))

#################################################
class ImageUploadParser(FileUploadParser):
    media_type = 'image/*'

#################################################
def build_excel_response(field_names,data,filename):
    try:
        wb = Workbook(write_only=True) 
        ws = wb.create_sheet()
        ws.append(field_names)
        for item in data:
            ws.append(item.values())
        
        s = StringIO()
        wb.save(s)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = u'attachment; filename={}.xlsx'.format(filename)
        response['Content-Length'] = s.len 
        response.write(s.getvalue())
        s.close()
        
        return response
    except Exception, exc:
        return build_exception_response()
